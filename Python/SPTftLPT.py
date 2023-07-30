# -*- coding: utf-8 -*-
"""
Created on Fri Aug 21 22:46:17 2020

@author: Administrator
"""


import pandas as pd
import openpyxl as op

import time
t = 10
start = time.time()
while t <=120:
    
    nan = 1 
    while nan < 2:
        '''input = pd.read_excel('Input'+str(nan)+'.xlsx')
    
        NoJob = len(input.CodeMat)
    
        Sfile = op.load_workbook('Input'+str(nan)+'.xlsx')
        Ssheet = Sfile.get_sheet_by_name('Sheet2')
    
        Ssheet.cell(row= 1 ,column=1).value = "Machine"
        Ssheet.cell(row= 1 ,column=2).value = "Size"
        Ssheet.cell(row= 1 ,column=3).value = "CodeMat"
        Ssheet.cell(row= 1 ,column=4).value = "thickness"
        Ssheet.cell(row= 1 ,column=5).value = "quantity"
        Ssheet.cell(row= 1 ,column=6).value = "Duedate"
        Ssheet.cell(row= 1 ,column=7).value = "ProcessingTime"
    
        k = 0
    
        for i in range(NoJob):
        
            for j in range(6,16):
                if input.iloc[i,j] > 0:
                    Ssheet.cell(row= k+2 ,column=1).value = j-5
                    Ssheet.cell(row= k+2 ,column=2).value = input.iloc[i,1]
                    Ssheet.cell(row= k+2 ,column=3).value = input.iloc[i,2]
                    Ssheet.cell(row= k+2 ,column=4).value = input.iloc[i,3]
                    Ssheet.cell(row= k+2 ,column=5).value = input.iloc[i,4]
                    Ssheet.cell(row= k+2 ,column=6).value = input.iloc[i,5]
                    Ssheet.cell(row= k+2 ,column=7).value = input.iloc[i,j]
                    k+=1
        Sfile.save('Input'+str(nan)+'.xlsx')'''
    
    
    #--------------------------------Rules---------------------------------
    
        CurrentMachineProcess = [0 for i in range(10)]
        CurrentCompletionTime = [0 for i in range(10)]
    
        FirstCT = [0 for i in range(10)]
        FirstJob = [0 for i in range(10)]
     
        ProT = [0 for i in range(10)]
        SetUp = [0 for i in range(10)]
    
        CountTardyJob = 0
        SumTardyJob = 0
    
    
    
        Data = pd.read_excel('Input.xlsx',"Sheet2")
        Data1 = Data
    
        #Add Column Update
        Data['UpdateProcessingTime'] = Data.ProcessingTime
        Data['Setup'] = Data.UpdateProcessingTime - Data.ProcessingTime
        #Data = Data[(Data.Machine!=3) & (Data.Machine!=4)]
    
        NoJob = len(list(dict.fromkeys(Data.CodeMat)))
    
        Afile = op.load_workbook('SPTftLPT.xlsx')
        Asheet = Afile.get_sheet_by_name('SPT-LPT-'+str(t))
        aab = 0 
        job = 1
    
        #While loop for schedule 
        while job <= NoJob:
            if max(CurrentCompletionTime)<t:
                indexmin = min(Data['UpdateProcessingTime'])
        
            #Find material code with min update processing time
                MatCode = Data[Data.UpdateProcessingTime==indexmin].CodeMat.values
         
                print("Matcode : ",list(dict.fromkeys(MatCode)))
                print("MatcodeChoose : ",list(dict.fromkeys(MatCode))[0])
            #Find Mahine
                MachineUse = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.UpdateProcessingTime==indexmin) ].Machine.values
                MachineAvaiable = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Machine.values
            #Specify size name
                SizeName = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.UpdateProcessingTime==indexmin) ].Size.values[0])
            #Specify thickness
                #Thickness = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.UpdateProcessingTime==indexmin) ].thickness.values[0])
                DueDate = int(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.UpdateProcessingTime==indexmin) ].Duedate.values[0])
                print("Machine Available : ", MachineAvaiable )   
        
                aa = list(MachineAvaiable)
       
                aa1 = [x for x in aa if x != int(MachineUse[0])]
        
            #aa = aa.remove(int(MachineUse))
        
                print("Before ",CurrentCompletionTime)
                SelectMachine = int(MachineUse[0])
        
            #Specify machine
                Index1 = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==int(MachineUse[0])) ].index.values
                TotalCompletion = CurrentCompletionTime[int(MachineUse[0])-1] + Data.UpdateProcessingTime[Index1].values[0]
                print(int(MachineUse[0]))
                print("Present : ", CurrentCompletionTime[int(MachineUse[0])-1])
                print("Setup : ",Data.UpdateProcessingTime[Index1].values[0])
        
            #Choosing machine
                if len(aa1) >=1:
         
                    print("Do",CurrentCompletionTime[int(MachineUse[0])-1])
                    print(aa)
            
            
                    for i in aa:
                        print("Total : ", TotalCompletion)
                        print("aa : ",i)
                    # print("Total Completion : ",float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
                        if CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])  < TotalCompletion:
                            print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]) ," < ", TotalCompletion)
                    
                            TotalCompletion = CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])
                            SelectMachine = i 
                        print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
            
    
    
        
                if FirstCT[SelectMachine - 1] <= 0:
                    FirstCT[SelectMachine - 1] = TotalCompletion
                    FirstJob[SelectMachine - 1] = list(dict.fromkeys(MatCode))[0]
                    print("First ", FirstCT[SelectMachine - 1])
       
        
                if TotalCompletion <= DueDate:
                    Tardiness = 0
                else:
                    Tardiness = TotalCompletion - DueDate
                    CountTardyJob = CountTardyJob +1
                    SumTardyJob = SumTardyJob + Tardiness
          
        
        
        
                Index = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==SelectMachine) ].index.values
                Setuptime = Data.UpdateProcessingTime[Index].values[0] - Data.ProcessingTime[Index].values[0]
                ProcessingT = Data.ProcessingTime[Index].values[0]
                DueD = Data.Duedate[Index].values[0]
         
        
                if (SetUp[SelectMachine-1]>=8 and ProT[SelectMachine-1]<8 and Setuptime >=8 ):
                    DataDe = Data[Data.Machine == SelectMachine]
                    indexmin = min(DataDe['Setup'])
                    indexm = min(Data[Data.Setup==indexmin].ProcessingTime)
                    MatCode = Data[(Data.Setup==indexmin) & (Data.Machine == SelectMachine )& (Data.ProcessingTime == indexm)].CodeMat.values
            
                    MachineUse = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Machine.values
                    MachineAvaiable = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Machine.values
                #Specify size name
                    SizeName = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Size.values[0])
               #Specify thickness
                    #Thickness = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].thickness.values[0])
                    DueDate = int(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Duedate.values[0])
                    print("Machine Available : ", MachineAvaiable )   
            
                    aa = list(MachineAvaiable)
            
                    aa1 = [x for x in aa if x != int(MachineUse[0])]
                
                #aa = aa.remove(int(MachineUse))
        
                    print("Before ",CurrentCompletionTime)
                    SelectMachine = int(MachineUse[0])
        
                #Specify machine
                    Index1 = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==int(MachineUse[0])) ].index.values
                    TotalCompletion = CurrentCompletionTime[int(MachineUse[0])-1] + Data.UpdateProcessingTime[Index1].values[0]
                    print(int(MachineUse[0]))
                    print("Present : ", CurrentCompletionTime[int(MachineUse[0])-1])
                    print("Setup : ",Data.UpdateProcessingTime[Index1].values[0])
        
                #Choosing machine
                    if len(aa1) >=1:
         
                        print("Do",CurrentCompletionTime[int(MachineUse[0])-1])
                        print(aa)
            
            
                        for i in aa:
                            print("Total : ", TotalCompletion)
                            print("aa : ",i)
                       # print("Total Completion : ",float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
                            if CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])  < TotalCompletion:
                                print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]) ," < ", TotalCompletion)
                                TotalCompletion = CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])
                                SelectMachine = i 
                            print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
            
        
                    if TotalCompletion <= DueDate:
                        Tardiness = 0
                    else:
                        Tardiness = TotalCompletion - DueDate
                        CountTardyJob = CountTardyJob +1
                        SumTardyJob = SumTardyJob + Tardiness
          
        
        
        
                    Index = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==SelectMachine) ].index.values
                    Setuptime = Data.UpdateProcessingTime[Index].values[0] - Data.ProcessingTime[Index].values[0]
                    ProcessingT = Data.ProcessingTime[Index].values[0]
                    DueD = Data.Duedate[Index].values[0]
            
           
        
        
                CurrentMachineProcess[SelectMachine-1] = str(Data[(Data.Machine == SelectMachine ) & (Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Size.values)[1:-1].replace("'","")
                CurrentCompletionTime[SelectMachine-1] = CurrentCompletionTime[SelectMachine-1] + float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])&(Data.Machine == SelectMachine)].UpdateProcessingTime.values[0])
        
            else:
                indexmin = max(Data['ProcessingTime'])
        
            #Find material code with min update processing time
                MatCode = Data[Data.ProcessingTime==indexmin].CodeMat.values
         
                print("Matcode : ",list(dict.fromkeys(MatCode)))
                print("MatcodeChoose : ",list(dict.fromkeys(MatCode))[0])
            #Find Mahine
                MachineUse = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.ProcessingTime==indexmin) ].Machine.values
                MachineAvaiable = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Machine.values
            #Specify size name
                SizeName = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.ProcessingTime==indexmin) ].Size.values[0])
            #Specify thickness
                #Thickness = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.ProcessingTime==indexmin) ].thickness.values[0])
                DueDate = int(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.ProcessingTime==indexmin) ].Duedate.values[0])
                print("Machine Available : ", MachineAvaiable )   
        
                aa = list(MachineAvaiable)
       
                aa1 = [x for x in aa if x != int(MachineUse[0])]
        
            #aa = aa.remove(int(MachineUse))
        
                print("Before ",CurrentCompletionTime)
                SelectMachine = int(MachineUse[0])
        
            #Specify machine
                Index1 = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==int(MachineUse[0])) ].index.values
                TotalCompletion = CurrentCompletionTime[int(MachineUse[0])-1] + Data.UpdateProcessingTime[Index1].values[0]
                print(int(MachineUse[0]))
                print("Present : ", CurrentCompletionTime[int(MachineUse[0])-1])
                print("Setup : ",Data.UpdateProcessingTime[Index1].values[0])
        
            #Choosing machine
                if len(aa1) >=1:
         
                    print("Do",CurrentCompletionTime[int(MachineUse[0])-1])
                    print(aa)
            
            
                    for i in aa:
                        print("Total : ", TotalCompletion)
                        print("aa : ",i)
                    # print("Total Completion : ",float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
                        if CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])  < TotalCompletion:
                            print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]) ," < ", TotalCompletion)
                    
                            TotalCompletion = CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])
                            SelectMachine = i 
                        print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
            
    
    
        
                if FirstCT[SelectMachine - 1] <= 0:
                    FirstCT[SelectMachine - 1] = TotalCompletion
                    FirstJob[SelectMachine - 1] = list(dict.fromkeys(MatCode))[0]
                    print("First ", FirstCT[SelectMachine - 1])
       
        
                if TotalCompletion <= DueDate:
                    Tardiness = 0
                else:
                    Tardiness = TotalCompletion - DueDate
                    CountTardyJob = CountTardyJob +1
                    SumTardyJob = SumTardyJob + Tardiness
          
        
        
        
                Index = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==SelectMachine) ].index.values
                Setuptime = Data.UpdateProcessingTime[Index].values[0] - Data.ProcessingTime[Index].values[0]
                ProcessingT = Data.ProcessingTime[Index].values[0]
                DueD = Data.Duedate[Index].values[0]
         
        
                if (SetUp[SelectMachine-1]>=8 and ProT[SelectMachine-1]<8 and Setuptime >=8 ):
                    DataDe = Data[Data.Machine == SelectMachine]
                    indexmin = min(DataDe['Setup'])
                    indexm = max(Data[Data.Setup==indexmin].ProcessingTime)
                    MatCode = Data[(Data.Setup==indexmin) & (Data.Machine == SelectMachine )& (Data.ProcessingTime == indexm)].CodeMat.values
            
                    MachineUse = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Machine.values
                    MachineAvaiable = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Machine.values
                #Specify size name
                    SizeName = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Size.values[0])
               #Specify thickness
                    #Thickness = str(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].thickness.values[0])
                    DueDate = int(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Setup==indexmin) ].Duedate.values[0])
                    print("Machine Available : ", MachineAvaiable )   
            
                    aa = list(MachineAvaiable)
            
                    aa1 = [x for x in aa if x != int(MachineUse[0])]
                
                #aa = aa.remove(int(MachineUse))
        
                    print("Before ",CurrentCompletionTime)
                    SelectMachine = int(MachineUse[0])
        
                #Specify machine
                    Index1 = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==int(MachineUse[0])) ].index.values
                    TotalCompletion = CurrentCompletionTime[int(MachineUse[0])-1] + Data.UpdateProcessingTime[Index1].values[0]
                    print(int(MachineUse[0]))
                    print("Present : ", CurrentCompletionTime[int(MachineUse[0])-1])
                    print("Setup : ",Data.UpdateProcessingTime[Index1].values[0])
        
                #Choosing machine
                    if len(aa1) >=1:
         
                        print("Do",CurrentCompletionTime[int(MachineUse[0])-1])
                        print(aa)
            
            
                        for i in aa:
                            print("Total : ", TotalCompletion)
                            print("aa : ",i)
                       # print("Total Completion : ",float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
                            if CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])  < TotalCompletion:
                                print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]) ," < ", TotalCompletion)
                                TotalCompletion = CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0])
                                SelectMachine = i 
                            print("Current : ", CurrentCompletionTime[i-1]+float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==i) ].UpdateProcessingTime.values[0]))
            
        
                    if TotalCompletion <= DueDate:
                        Tardiness = 0
                    else:
                        Tardiness = TotalCompletion - DueDate
                        CountTardyJob = CountTardyJob +1
                        SumTardyJob = SumTardyJob + Tardiness
          
        
        
        
                    Index = Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0]) & (Data.Machine==SelectMachine) ].index.values
                    Setuptime = Data.UpdateProcessingTime[Index].values[0] - Data.ProcessingTime[Index].values[0]
                    ProcessingT = Data.ProcessingTime[Index].values[0]
                    DueD = Data.Duedate[Index].values[0]
            
           
        
        
                CurrentMachineProcess[SelectMachine-1] = str(Data[(Data.Machine == SelectMachine ) & (Data.CodeMat == list(dict.fromkeys(MatCode))[0])].Size.values)[1:-1].replace("'","")
                CurrentCompletionTime[SelectMachine-1] = CurrentCompletionTime[SelectMachine-1] + float(Data[(Data.CodeMat == list(dict.fromkeys(MatCode))[0])&(Data.Machine == SelectMachine)].UpdateProcessingTime.values[0])
            
            
            Asheet.cell(row= 1 ,column=4).value = "Codemat"
            Asheet.cell(row= job+1 ,column=4).value = list(dict.fromkeys(MatCode))[0]
        
            Asheet.cell(row= 1 ,column=2).value = "Size"
            Asheet.cell(row= job+1 ,column=2).value = SizeName
        
            #Asheet.cell(row= 1 ,column=3).value = "Thickness"
            #Asheet.cell(row= job+1 ,column=3).value = Thickness
        
            Asheet.cell(row= 1 ,column=1).value = "Machine"
            Asheet.cell(row= job+1 ,column=1).value = SelectMachine
        
            Asheet.cell(row= 1 ,column=5).value = "Setup Time"
            Asheet.cell(row= job+1 ,column=5).value = Data.UpdateProcessingTime[Index].values[0] - Data.ProcessingTime[Index].values[0]
        
            Asheet.cell(row= 1 ,column=7).value = "CurrentCompletionTime"
            Asheet.cell(row= job+1 ,column=7).value = CurrentCompletionTime[SelectMachine-1]
        
            Asheet.cell(row= 1 ,column=6).value = "Processingtime"
            Asheet.cell(row= job+1 ,column=6).value = ProcessingT
        
            Asheet.cell(row= 1 ,column=8).value = "DueDate"
            Asheet.cell(row= job+1 ,column=8).value = DueD
        
            Asheet.cell(row= 1 ,column=9).value = "Tardiness"
            Asheet.cell(row= job+1 ,column=9).value = Tardiness
        
            Afile.save('SPTftLPT.xlsx')
        
        
            ProT [SelectMachine-1]= ProcessingT
            SetUp [SelectMachine-1]=Setuptime
        
        
        
        
            if SelectMachine ==1 :
            
                Setup = pd.read_excel('Setup.xlsx','TM01')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
            
            if SelectMachine == 2 :
                Setup = pd.read_excel('Setup.xlsx','TM02')
           
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
    
            if SelectMachine == 3 :
                Setup = pd.read_excel('Setup.xlsx','TM03')
           
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
            
            if SelectMachine == 4 :
                Setup = pd.read_excel('SetupTime.xlsx','TM03-04')
           
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    
              
            if SelectMachine == 5 :
                Setup = pd.read_excel('SetupTime.xlsx','TM05')
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
            if SelectMachine == 6 :
                Setup = pd.read_excel('SetupTime.xlsx','TM06')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
            if SelectMachine ==   7:
                Setup = pd.read_excel('SetupTime.xlsx','TM07')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
            if SelectMachine == 8 :
                Setup = pd.read_excel('SSetup.xlsx','Sheet3')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
            if SelectMachine == 9 :
                Setup = pd.read_excel('SSetup.xlsx','Sheet3')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
            if SelectMachine == 10 :
                Setup = pd.read_excel('SSetup.xlsx','Sheet3')
            
                for j in Data[Data.Machine == SelectMachine].index.values:
                    Data.UpdateProcessingTime[j] = Data.ProcessingTime[j] + float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
                    Data.Setup[j]= float(Setup[Setup.Size == CurrentMachineProcess[SelectMachine-1]][[str(Data[Data.index == j].Size.values)[1:-1].replace("'","")]].values[0])
        
        
            Data = Data[Data.CodeMat!= list(dict.fromkeys(MatCode))[0]]
       
            job +=1
        
        Bfile = op.load_workbook('SPTftLPT.xlsx')
        Bsheet = Bfile.get_sheet_by_name('compare')
    
        Bsheet.cell(row= 1 ,column=1).value = "Rule"
        Bsheet.cell(row= 1 ,column=2).value = "Machine 1"
        Bsheet.cell(row= 1 ,column=3).value = "Machine 2"
        Bsheet.cell(row= 1 ,column=4).value = "Machine 3"
        Bsheet.cell(row= 1 ,column=5).value = "Machine 4"
        Bsheet.cell(row= 1 ,column=6).value = "Machine 5"
        Bsheet.cell(row= 1 ,column=7).value = "Machine 6"
        Bsheet.cell(row= 1 ,column=8).value = "Machine 7"
        Bsheet.cell(row= 1 ,column=9).value = "Machine 8"
        Bsheet.cell(row= 1 ,column=10).value = "Machine 9"
        Bsheet.cell(row= 1 ,column=11).value = "Machine 10"
    
   
        Bsheet.cell(row= 1 ,column=12).value = "Makespan"
        Bsheet.cell(row= t ,column=12).value = max(CurrentCompletionTime)
    
        Bsheet.cell(row= 1 ,column=13).value = "Number of Tardy Job"
        Bsheet.cell(row= t ,column=13).value = CountTardyJob
    
        Bsheet.cell(row= 1 ,column=14).value = "Average Tardy Hours"
        Bsheet.cell(row= t ,column=14).value = SumTardyJob
  
        Bsheet.cell(row= t,column=1).value = "SPT-LPT"+str(t)
        Bsheet.cell(row= t+1 ,column=1).value = "Slack"    
    
    
        Bfile.save('SPTftLPT.xlsx')
        
        nan +=1
    t+=10

end = time.time()

print("Computation Time : ",end-start)